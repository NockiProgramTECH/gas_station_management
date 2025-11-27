"""
Exporteur Excel Avancé
Génère des rapports Excel formatés avec graphiques
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from database.db_manager import DatabaseManager
from analytics_engine import AnalyticsEngine
from datetime import datetime, timedelta
from typing import Dict, List
import os

class ExcelExporter:
    """
    Classe pour exporter des données vers Excel
    
    Attributes:
        db (DatabaseManager): Gestionnaire de base de données
        analytics (AnalyticsEngine): Moteur d'analytics
    """
    
    def __init__(self):
        """Initialise l'exporteur Excel"""
        self.db = DatabaseManager()
        self.analytics = AnalyticsEngine()
        
        # Styles
        self.header_font = Font(bold=True, size=12, color="FFFFFF")
        self.header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        self.title_font = Font(bold=True, size=16)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export_monthly_report(self, mois: str, annee: str, output_path: str) -> bool:
        """
        Exporte un rapport mensuel complet
        
        Args:
            mois (str): Mois (MM)
            annee (str): Année (YYYY)
            output_path (str): Chemin de sortie
            
        Returns:
            bool: True si succès
        """
        try:
            wb = Workbook()
            
            # Supprimer la feuille par défaut
            wb.remove(wb.active)
            
            # 1. Feuille Résumé
            self._create_summary_sheet(wb, mois, annee)
            
            # 2. Feuille Ventes Détaillées
            self._create_sales_sheet(wb, mois, annee)
            
            # 3. Feuille Performance Pompistes
            self._create_performance_sheet(wb, mois, annee)
            
            # 4. Feuille Inventaire
            self._create_inventory_sheet(wb, mois, annee)
            
            # 5. Feuille Pertes
            self._create_losses_sheet(wb, mois, annee)
            
            # Sauvegarder
            wb.save(output_path)
            return True
        
        except Exception as e:
            print(f"❌ Erreur export Excel: {e}")
            return False
    
    def _create_summary_sheet(self, wb: Workbook, mois: str, annee: str):
        """
        Crée la feuille de résumé
        
        Args:
            wb (Workbook): Classeur Excel
            mois (str): Mois
            annee (str): Année
        """
        ws = wb.create_sheet("Résumé", 0)
        
        # Titre
        ws['A1'] = "RAPPORT MENSUEL - STATION SERVICE TOTAL"
        ws['A1'].font = Font(bold=True, size=18, color="EE3124")
        ws.merge_cells('A1:F1')
        
        ws['A2'] = f"Période: {mois}/{annee}"
        ws['A2'].font = Font(size=12, italic=True)
        ws.merge_cells('A2:F2')
        
        ws['A3'] = f"Généré le: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A3'].font = Font(size=10, italic=True)
        ws.merge_cells('A3:F3')
        
        # Calculer les statistiques
        date_debut = f"{annee}-{mois}-01"
        if mois == '12':
            date_fin = f"{int(annee)+1}-01-01"
        else:
            date_fin = f"{annee}-{int(mois)+1:02d}-01"
        
        conn = self.db.get_connection()
        cursor = conn.cursor()
        
        # Ventes totales
        cursor.execute('''
            SELECT 
                COUNT(*) as nb_transactions,
                SUM(quantite_vendue) as total_litres,
                SUM(montant_total) as total_montant,
                AVG(montant_total) as vente_moyenne
            FROM ventes
            WHERE date >= ? AND date < ?
        ''', (date_debut, date_fin))
        
        stats = cursor.fetchone()
        
        # Ventes par type
        cursor.execute('''
            SELECT 
                r.type_carburant,
                SUM(v.quantite_vendue) as litres,
                SUM(v.montant_total) as montant,
                COUNT(*) as nb_ventes
            FROM ventes v
            JOIN reservoirs r ON v.reservoir_id = r.id
            WHERE v.date >= ? AND v.date < ?
            GROUP BY r.type_carburant
        ''', (date_debut, date_fin))
        
        ventes_type = cursor.fetchall()
        conn.close()
        
        # Section Statistiques Globales
        row = 5
        ws[f'A{row}'] = "📊 STATISTIQUES GLOBALES"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="1976D2")
        ws.merge_cells(f'A{row}:F{row}')
        
        row += 2
        headers = ['Indicateur', 'Valeur', 'Unité']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        data_stats = [
            ['Nombre de transactions', stats['nb_transactions'] if stats else 0, 'transactions'],
            ['Volume total vendu', f"{stats['total_litres']:,.0f}" if stats and stats['total_litres'] else '0', 'litres'],
            ['Chiffre d\'affaires', f"{stats['total_montant']:,.2f}" if stats and stats['total_montant'] else '0', 'FCFA'],
            ['Vente moyenne', f"{stats['vente_moyenne']:,.2f}" if stats and stats['vente_moyenne'] else '0', 'FCFA']
        ]
        
        for data_row in data_stats:
            row += 1
            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row, col, value)
                cell.border = self.border
                if col == 2:
                    cell.alignment = Alignment(horizontal='right')
        
        # Section Ventes par Type
        row += 3
        ws[f'A{row}'] = "⛽ VENTES PAR TYPE DE CARBURANT"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="1976D2")
        ws.merge_cells(f'A{row}:F{row}')
        
        row += 2
        headers = ['Type', 'Nombre', 'Volume (L)', 'Montant (FCFA)', 'Prix Moyen']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        for vente in ventes_type:
            row += 1
            prix_moyen = vente['montant'] / vente['litres'] if vente['litres'] > 0 else 0
            
            data = [
                vente['type_carburant'],
                vente['nb_ventes'],
                f"{vente['litres']:,.0f}",
                f"{vente['montant']:,.2f}",
                f"{prix_moyen:.2f}"
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row, col, value)
                cell.border = self.border
                if col > 1:
                    cell.alignment = Alignment(horizontal='right')
        
        # Graphique
        if len(ventes_type) > 0:
            chart = PieChart()
            chart.title = "Répartition des Ventes par Type"
            chart.style = 10
            
            labels = Reference(ws, min_col=1, min_row=row-len(ventes_type)+1, max_row=row)
            data_ref = Reference(ws, min_col=4, min_row=row-len(ventes_type), max_row=row)
            
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(labels)
            
            ws.add_chart(chart, f"H5")
        
        # Ajuster largeurs
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
    
    def _create_sales_sheet(self, wb: Workbook, mois: str, annee: str):
        """
        Crée la feuille des ventes détaillées
        
        Args:
            wb (Workbook): Classeur
            mois (str): Mois
            annee (str): Année
        """
        ws = wb.create_sheet("Ventes Détaillées")
        
        # Titre
        ws['A1'] = "VENTES DÉTAILLÉES"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:H1')
        
        # Headers
        headers = ['Date', 'Période', 'Pompiste', 'Type Carburant', 
                  'Quantité (L)', 'Montant (FCFA)', 'Heure Début', 'Heure Fin']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(3, col, header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Données
        date_debut = f"{annee}-{mois}-01"
        if mois == '12':
            date_fin = f"{int(annee)+1}-01-01"
        else:
            date_fin = f"{annee}-{int(mois)+1:02d}-01"
        
        conn = self.db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT 
                v.date, v.periode, v.quantite_vendue, v.montant_total,
                v.heure_debut, v.heure_fin,
                p.prenom || ' ' || p.nom as pompiste,
                r.type_carburant
            FROM ventes v
            JOIN pompistes p ON v.pompiste_id = p.id
            JOIN reservoirs r ON v.reservoir_id = r.id
            WHERE v.date >= ? AND v.date < ?
            ORDER BY v.date, v.periode
        ''', (date_debut, date_fin))
        
        ventes = cursor.fetchall()
        conn.close()
        
        row = 4
        for vente in ventes:
            data = [
                vente['date'],
                vente['periode'].title(),
                vente['pompiste'],
                vente['type_carburant'],
                f"{vente['quantite_vendue']:,.2f}",
                f"{vente['montant_total']:,.2f}",
                vente['heure_debut'],
                vente['heure_fin']
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row, col, value)
                cell.border = self.border
                if col in [5, 6]:
                    cell.alignment = Alignment(horizontal='right')
            
            row += 1
        
        # Totaux
        if len(ventes) > 0:
            row += 1
            ws.cell(row, 4, "TOTAL:").font = Font(bold=True)
            ws.cell(row, 5, f"=SUM(E4:E{row-2})").font = Font(bold=True)
            ws.cell(row, 6, f"=SUM(F4:F{row-2})").font = Font(bold=True)
        
        # Ajuster largeurs
        for col in range(1, 9):
            ws.column_dimensions[chr(64+col)].width = 15
    
    def _create_performance_sheet(self, wb: Workbook, mois: str, annee: str):
        """
        Crée la feuille de performance des pompistes
        
        Args:
            wb (Workbook): Classeur
            mois (str): Mois
            annee (str): Année
        """
        ws = wb.create_sheet("Performance Pompistes")
        
        # Titre
        ws['A1'] = "PERFORMANCE DES POMPISTES"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:G1')
        
        # Récupérer les performances
        performances = self.analytics.analyser_performance_pompistes(mois, annee)
        
        # Headers
        headers = ['Rang', 'Badge', 'Pompiste', 'Nb Ventes', 'Volume (L)', 
                  'Montant (FCFA)', 'Score']
        
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Données
        for perf in performances:
            row += 1
            data = [
                perf['rang'],
                perf['badge'],
                perf['nom_complet'],
                perf['nb_ventes'],
                f"{perf['total_litres']:,.2f}",
                f"{perf['total_montant']:,.2f}",
                f"{perf['efficacite']:.1f}"
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row, col, value)
                cell.border = self.border
                if col > 3:
                    cell.alignment = Alignment(horizontal='right')
                
                # Colorer le top 3
                if perf['rang'] == 1:
                    cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                elif perf['rang'] == 2:
                    cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
                elif perf['rang'] == 3:
                    cell.fill = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")
        
        # Graphique
        if len(performances) > 0:
            chart = BarChart()
            chart.title = "Performance par Pompiste"
            chart.style = 10
            chart.y_axis.title = "Montant (FCFA)"
            
            data_ref = Reference(ws, min_col=6, min_row=3, max_row=row)
            cats = Reference(ws, min_col=3, min_row=4, max_row=row)
            
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            
            ws.add_chart(chart, "I3")
        
        # Ajuster largeurs
        for col in range(1, 8):
            ws.column_dimensions[chr(64+col)].width = 18
    
    def _create_inventory_sheet(self, wb: Workbook, mois: str, annee: str):
        """
        Crée la feuille d'inventaire
        
        Args:
            wb (Workbook): Classeur
            mois (str): Mois
            annee (str): Année
        """
        ws = wb.create_sheet("Inventaire")
        
        ws['A1'] = "ÉTAT DE L'INVENTAIRE"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        
        # Récupérer les réservoirs
        reservoirs = self.db.obtenir_reservoirs()
        today = datetime.now().strftime('%Y-%m-%d')
        
        # Headers
        headers = ['Réservoir', 'Type', 'Capacité Max (L)', 'Niveau Actuel (L)', 
                  'Niveau (%)', 'Statut']
        
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Données
        for reservoir in reservoirs:
            row += 1
            niveau = self.db.obtenir_niveau_quotidien(reservoir['id'], today)
            
            if niveau:
                quantite = niveau['quantite_debut'] + niveau['quantite_entree']
                pct = (quantite / reservoir['capacite_max']) * 100
            else:
                quantite = 0
                pct = 0
            
            # Déterminer statut
            if pct < 20:
                statut = "CRITIQUE"
                fill_color = "F44336"
            elif pct < 50:
                statut = "BAS"
                fill_color = "FF9800"
            else:
                statut = "BON"
                fill_color = "4CAF50"
            
            data = [
                reservoir['nom'],
                reservoir['type_carburant'],
                f"{reservoir['capacite_max']:,.0f}",
                f"{quantite:,.0f}",
                f"{pct:.1f}%",
                statut
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row, col, value)
                cell.border = self.border
                if col in [3, 4]:
                    cell.alignment = Alignment(horizontal='right')
                elif col == 5:
                    cell.alignment = Alignment(horizontal='center')
                elif col == 6:
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
        
        # Ajuster largeurs
        for col in range(1, 7):
            ws.column_dimensions[chr(64+col)].width = 20
    
    def _create_losses_sheet(self, wb: Workbook, mois: str, annee: str):
        """
        Crée la feuille des pertes
        
        Args:
            wb (Workbook): Classeur
            mois (str): Mois
            annee (str): Année
        """
        ws = wb.create_sheet("Analyse des Pertes")
        
        ws['A1'] = "ANALYSE DES PERTES"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:G1')
        
        # Calculer pertes du mois
        date_debut = f"{annee}-{mois}-01"
        if mois == '12':
            date_fin_dt = datetime(int(annee)+1, 1, 1)
        else:
            date_fin_dt = datetime(int(annee), int(mois)+1, 1)
        
        date_fin = date_fin_dt.strftime('%Y-%m-%d')
        
        # Headers
        headers = ['Date', 'Réservoir', 'Type', 'Stock Début', 'Vendu', 
                  'Perte (L)', 'Perte (FCFA)']
        
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Parcourir chaque jour
        current_date = datetime.strptime(date_debut, '%Y-%m-%d')
        end_date = datetime.strptime(date_fin, '%Y-%m-%d')
        
        total_pertes_litres = 0
        total_pertes_fcfa = 0
        
        while current_date < end_date:
            date_str = current_date.strftime('%Y-%m-%d')
            pertes = self.analytics.calculer_pertes_automatiques(date_str)
            
            for perte_res in pertes['reservoirs']:
                if perte_res['perte_litres'] > 0:
                    row += 1
                    data = [
                        date_str,
                        perte_res['nom'],
                        perte_res['type'],
                        f"{perte_res['niveau_actuel']:,.0f}",
                        "N/A",
                        f"{perte_res['perte_litres']:,.2f}",
                        f"{perte_res['perte_argent']:,.2f}"
                    ]
                    
                    for col, value in enumerate(data, 1):
                        cell = ws.cell(row, col, value)
                        cell.border = self.border
                        if col > 3:
                            cell.alignment = Alignment(horizontal='right')
                    
                    total_pertes_litres += perte_res['perte_litres']
                    total_pertes_fcfa += perte_res['perte_argent']
            
            current_date += timedelta(days=1)
        
        # Totaux
        if row > 3:
            row += 2
            ws.cell(row, 5, "TOTAL PERTES:").font = Font(bold=True, size=12)
            ws.cell(row, 6, f"{total_pertes_litres:,.2f} L").font = Font(bold=True, size=12)
            ws.cell(row, 7, f"{total_pertes_fcfa:,.2f} FCFA").font = Font(bold=True, size=12)
        
        # Ajuster largeurs
        for col in range(1, 8):
            ws.column_dimensions[chr(64+col)].width = 18
    
    def export_daily_sales(self, date: str, output_path: str) -> bool:
        """
        Exporte les ventes d'une journée
        
        Args:
            date (str): Date (YYYY-MM-DD)
            output_path (str): Chemin de sortie
            
        Returns:
            bool: True si succès
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = f"Ventes {date}"
            
            # Titre
            ws['A1'] = f"VENTES DU {date}"
            ws['A1'].font = Font(bold=True, size=16)
            ws.merge_cells('A1:H1')
            
            # Headers
            headers = ['Période', 'Pompiste', 'Type Carburant', 'Quantité (L)', 
                      'Montant (FCFA)', 'Heure Début', 'Heure Fin', 'Remarques']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(3, col, header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = self.border
            
            # Données
            ventes = self.db.obtenir_ventes_par_date(date)
            
            row = 4
            for vente in ventes:
                data = [
                    vente['periode'].title(),
                    f"{vente['prenom']} {vente['nom']}",
                    vente['type_carburant'],
                    f"{vente['quantite_vendue']:,.2f}",
                    f"{vente['montant_total']:,.2f}",
                    vente['heure_debut'],
                    vente['heure_fin'],
                    ""
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row, col, value)
                    cell.border = self.border
                    if col in [4, 5]:
                        cell.alignment = Alignment(horizontal='right')
                
                row += 1
            
            # Totaux
            if len(ventes) > 0:
                row += 1
                ws.cell(row, 3, "TOTAL:").font = Font(bold=True)
                ws.cell(row, 4, f"=SUM(D4:D{row-2})").font = Font(bold=True)
                ws.cell(row, 5, f"=SUM(E4:E{row-2})").font = Font(bold=True)
            
            # Ajuster largeurs
            for col in range(1, 9):
                ws.column_dimensions[chr(64+col)].width = 15
            
            wb.save(output_path)
            return True
        
        except Exception as e:
            print(f"❌ Erreur export: {e}")
            return False